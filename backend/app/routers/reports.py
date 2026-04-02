from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session
from sqlalchemy import extract, func
from datetime import date
from typing import Optional
from app.database import get_db
from app import models
from app.auth import get_current_active_user
import io

router = APIRouter(prefix="/api/reports", tags=["Reports"])

def _get_monthly_data(db: Session, year: int, zone_id: Optional[int] = None):
    rows = []
    for month in range(1, 13):
        q = (
            db.query(
                func.sum(models.Bill.amount_billed).label("billed"),
                func.sum(models.Bill.amount_paid).label("collected"),
                func.sum(models.Bill.units_consumed).label("consumption"),
                func.sum(models.Bill.nrw_loss).label("nrw"),
                func.count(models.Bill.id).label("bills"),
            )
            .join(models.Customer)
            .filter(
                extract("year", models.Bill.bill_date) == year,
                extract("month", models.Bill.bill_date) == month,
            )
        )
        if zone_id:
            q = q.filter(models.Customer.zone_id == zone_id)
        r = q.first()
        billed = r.billed or 0
        collected = r.collected or 0
        rows.append({
            "month": date(year, month, 1).strftime("%B %Y"),
            "billed": round(billed, 2),
            "collected": round(collected, 2),
            "collection_rate": round((collected / billed * 100) if billed else 0, 1),
            "consumption_m3": round(r.consumption or 0, 2),
            "nrw_m3": round(r.nrw or 0, 2),
            "bill_count": r.bills or 0,
        })
    return rows

@router.get("/excel")
def export_excel(
    year: Optional[int] = None,
    zone_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_active_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    yr = year or date.today().year
    rows = _get_monthly_data(db, yr, zone_id)

    wb = Workbook()
    ws = wb.active
    ws.title = f"KIWASCO Report {yr}"

    # === Header branding ===
    ws.merge_cells("A1:G1")
    ws["A1"] = "KIWASCO — Kisumu Water & Sewerage Company"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1A56DB")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    zone_label = f"Zone Filter: Zone ID {zone_id}" if zone_id else "All Zones"
    ws["A2"] = f"Annual Revenue & Billing Report — {yr} | {zone_label}"
    ws["A2"].font = Font(size=11, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    # === Column headers ===
    headers = ["Month", "Billed (KES)", "Collected (KES)", "Collection Rate (%)",
               "Consumption (m³)", "NRW (m³)", "No. of Bills"]
    header_fill = PatternFill("solid", fgColor="1A56DB")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # === Data rows ===
    for row_num, row in enumerate(rows, start=5):
        values = [
            row["month"], row["billed"], row["collected"],
            row["collection_rate"], row["consumption_m3"],
            row["nrw_m3"], row["bill_count"],
        ]
        fill = PatternFill("solid", fgColor="EBF5FB") if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = border
            cell.fill = fill
            if col in [2, 3]:
                cell.number_format = '#,##0.00'
            if col == 4:
                cell.number_format = '0.0"%"'

    # Column widths
    widths = [18, 18, 18, 20, 20, 15, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary row
    sum_row = len(rows) + 5
    ws.cell(row=sum_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in [2, 3, 5, 6]:
        ws.cell(row=sum_row, column=col, value=f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{sum_row-1})")
        ws.cell(row=sum_row, column=col).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"KIWASCO_Report_{yr}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/summary")
def report_summary(
    year: Optional[int] = None,
    zone_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_active_user),
):
    yr = year or date.today().year
    return _get_monthly_data(db, yr, zone_id)
